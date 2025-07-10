import csv
import json
import pandas as pd
from typing import List, Type
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import date



def convert_dates(obj):
    if isinstance(obj, date):
        return obj.strftime("%Y-%m-%d")
    raise TypeError("Type not serializable")



def save_csv(data: List[object], filename: str, fields: List[str]):
    with open(filename, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(fields)
        for obj in data:
            writer.writerow([getattr(obj, field) for field in fields])



def save_json(data: List[object], filename: str, fields: List[str]):
    with open(filename, "w") as f:
        json.dump(
            [{field: convert_dates(getattr(obj, field)) if isinstance(getattr(obj, field), date) else getattr(obj, field)
              for field in fields} for obj in data], f, indent=4)



def save_xlsx(data: List[object], sheet_name: str, filename: str, fields: List[str]):
    df = pd.DataFrame([{field: getattr(obj, field) for field in fields} for obj in data])


    try:
        workbook = load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()


    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]


    sheet = workbook.create_sheet(title=sheet_name)


    for col_num, field in enumerate(fields, 1):
        cell = sheet.cell(row=1, column=col_num, value=field)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )


    for row_num, row_data in enumerate(df.values, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for col in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        col_letter = col[0].column_letter  # Pl. A, B, C...
        sheet.column_dimensions[col_letter].width = max_length + 2


    workbook.save(filename)